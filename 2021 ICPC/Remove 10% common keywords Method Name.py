import openpyxl

z = "Com_Met_5_per.xlsx"
#x% 19767  x*19767/100

support = 988
i = 0
x = {}
y = []
listDB = []
ListCommon = []
row_no = 2

# Read the excell file and sheet for input
wb = openpyxl.load_workbook('All in one nlt3.xlsx')
sheet_input = wb['Sheet1']

# Read the excell file and sheet for output
wb2 = openpyxl.Workbook()
sheet_output = wb2.active


for row in sheet_input['d2':'d19769']:  # read perticular cell not all cell in the sheet
    try:
        for cell in row:
            x[i] = cell.value  # put the cell value into the list x[i]
        x[i] = sorted(x[i].split())  # split the value into words
        listDB.append(x[i])
        i = i + 1
    except AttributeError:
        x[i] = " "
        i = i + 1

number_of_cell = i

for j in range(0, number_of_cell):
    try:
        y = (set(y).union(set(x[j])))
    except TypeError:
        y=set(y).union(set(y))

count = [0] * len(y)  # All the word untill lenth of y is set their count to 0
wordCountDict = dict(zip(list(y), count))

for key in wordCountDict:
    for db in listDB:
        if key in db:
             wordCountDict[key] += 1

for key in wordCountDict.keys():
        if (wordCountDict.get(key) >= support):
            ListCommon.append(key)
#print("Word with greter than support:", ListCommon)

for k in range(0, number_of_cell):
    try:
        x[k] = sorted(set(x[k]).difference(set(ListCommon)))
        x[k] = str(x[k])
        x[k] = x[k].replace(',', '')
        x[k] = x[k].replace('\'', '')
        x[k] = x[k].replace(']', '')
        x[k] = x[k].replace('[', '')
        a2 = sheet_output.cell(row=row_no, column=3)
        a2.value = x[k]
        #print(x[k])
        row_no = row_no + 1
        wb2.save(z)
    except TypeError:
        a2 = sheet_output.cell(row=row_no, column=3)
        a2.value = " "
        #print(" ")
        row_no = row_no + 1
        wb2.save(z)
wb2.save(z)