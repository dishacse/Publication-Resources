import re
import openpyxl
import spacy
import re

# Initialize spacy 'en' model, keeping only tagger component needed for lemmatization
nlp = spacy.load('en_core_web_sm', disable=['parser', 'ner'])

number_of_cell = 19770
row_no = 1
i = 0
x = {}

f = open('output13.txt', 'w+')
z = "nlt3method.xlsx"

# Read the excell file and sheet for input
wb = openpyxl.load_workbook('All Method Name.xlsx')
sheet_input = wb['Sheet']

# Read the excell file and sheet for output
wb2 = openpyxl.Workbook()
sheet_output = wb2.active

for row in sheet_input['B1':'B19770']:  # read perticular cell not all cell in the sheet
    try:
        for cell in row:
            x[i] = cell.value  # put the cell value into the list x[i]
            word = str(x[i])
            sentence = word.lower()

            # Parse the sentence using the loaded 'en' model object `nlp`
            doc = nlp(sentence)
            print(sentence)
            # Extract the lemma for each token and join
            data = " ".join([token.lemma_ for token in doc])
            data = data.split()
            data = sorted(set(data))
            data = str(data)
            data = data.replace('[', '')
            data = data.replace(']', '')
            data = data.replace('\'', '')
            data = data.replace(',', '')
            data = data.replace('-PRON- ', '')
            print(data)

            word = re.sub(r'\b\w{1,3} \b', '', data)  # remove any word less than three charecter
            word = re.sub(r'\b\w{1,3}\b', '', word)  # remove any word less than three charecter

            print(word)
            print("\n\n")

            a2 = sheet_output.cell(row=row_no, column=2)
            a2.value = word
            row_no = row_no + 1

            wb2.save(z)
        wb2.save(z)
    except:
         row_no = row_no + 1
         pass