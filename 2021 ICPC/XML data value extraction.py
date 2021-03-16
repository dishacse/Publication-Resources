///
Author: Md. Shamsujjoha, Prof. John Grundy, Dr. Li Li, Dr. Qinghua Lu and Dr. Hourieh Khalajzade
Python Version: 3.7 or later
///

import os
import fnmatch
import openpyxl
import glob
import re
import warnings
warnings.filterwarnings('ignore')

wb2 = openpyxl.Workbook()
sheet_output = wb2.active
z = "XML_1.xlsx" #
row_no = 1

myFiles = ['001997E04F504112B5E018DD86261CFE47884E3EE1CEABB14AFCB4B8C122C6AB'] // this is an example app, during the computation we have use 5000 apps at a time where SHA256 are seperated by commas

for filename in myFiles:
    f5 = open('out1.txt', 'w+', encoding="utf8") #

    a2 = sheet_output.cell(row=row_no, column=1)
    y = str(filename)
    a2.value = y
    wb2.save(z)

    f_XML = open('out2.txt', 'w+', encoding="utf8") #
    folder_path= os.path.join('D1', filename, 'res', 'values' ) #
    for dirpath, dirs, files in os.walk(folder_path):
        for filename in fnmatch.filter(files, 'strings.xml'):
            try:
                i=0
                f=open(os.path.join(dirpath, filename), encoding='ISO-8859-1')
                for word in f:
                    string_with_quotes = word
                    Find_double_quotes = re.compile('">([^>]*)</',
                                                    re.DOTALL | re.MULTILINE | re.IGNORECASE)  # Ignore case not needed here, but can be useful.
                    list_of_quotes = Find_double_quotes.findall(string_with_quotes)
                    list_of_quotes = str (list_of_quotes)

                    list_of_quotes = re.sub(r'[^A-Za-z\s]+', "", list_of_quotes)  # Remove any single letter
                    list_of_quotes = re.sub(r"\b[a-zA-Z]\b", "", list_of_quotes)  # Remove any single letter

                    print(list_of_quotes, end = " ")
                    print(list_of_quotes, end = " ", file = f_XML)
            except:
                pass
    f_XML.close()

    f6 = open('out2.txt', 'r+', encoding="utf8") #
    a = sheet_output.cell(row=row_no, column=2)
    x = f6.read()
    a.value = x
    wb2.save(z)
    f6.close()
    row_no = row_no + 1
    f5.close()
wb2.save(z)