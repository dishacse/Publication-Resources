///
Author: Md. Shamsujjoha, Prof. John Grundy, Dr. Li Li, Dr. Qinghua Lu and Dr. Hourieh Khalajzade
Python Version: 3.7 or later
///

import pytesseract
import openpyxl
from PIL import Image
import re
import os
import fnmatch
import glob
import warnings
warnings.filterwarnings('ignore')

wb2 = openpyxl.Workbook()
sheet_output = wb2.active
z = "GUI_1.xlsx"
row_no = 1
column_no = 2

myFiles = ['001997E04F504112B5E018DD86261CFE47884E3EE1CEABB14AFCB4B8C122C6AB'] // this is an example app, during the computation we have use 5000 apps at a time where SHA256 are seperated by commas

for filename in myFiles:

    a2 = sheet_output.cell(row=row_no, column=1)
    y = str(filename)
    a2.value = y
    wb2.save(z)
    
    f2 = open('out1.txt', 'w+', encoding="utf8")
    folder_path= os.path.join('D1', filename)
    for dirpath, dirs, files in os.walk(folder_path):
        for filename in fnmatch.filter(files, '*.*'):
            try:
                f = Image.open(os.path.join(dirpath, filename))
                text = pytesseract.image_to_string(f, lang='eng')
                out_str = str(text)
                out_str = re.sub(r"\b[a-zA-Z]\b", "", out_str) #Remove any single letter
                out_str = re.sub(r'[^A-Za-z\s]+', "", out_str) #Remove any non alphabatic charecter
                out_str = out_str.replace('(', '')
                out_str = out_str.replace(')', '')
                out_str = out_str.replace(',', '')
                out_str = out_str.replace('.', '')
                out_str = out_str.replace('  ', ' ')
                print(out_str, end=" ", file=f2)
            except :
                pass
    f2.close()

    f3 = open('out1.txt', 'r+', encoding="utf8")
    a = sheet_output.cell(row=row_no, column=column_no)
    x = f3.read()
    print(x)
    a.value = x
    wb2.save(z)
    f3.close()
    row_no = row_no + 1    
wb2.save(z)