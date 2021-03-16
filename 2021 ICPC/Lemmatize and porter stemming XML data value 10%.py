import openpyxl
import spacy
from nltk.stem import PorterStemmer
from nltk.stem import LancasterStemmer

from nltk.tokenize import sent_tokenize, word_tokenize

def stemSentence(sentence):
    token_words=word_tokenize(sentence)
    token_words
    stem_sentence=[]
    for word in token_words:
        stem_sentence.append(porter.stem(word))
        stem_sentence.append(" ")
    return "".join(stem_sentence)


porter = PorterStemmer()
lancaster=LancasterStemmer()

row_no = 2
i = 0
x = {}

z = "Por_xml_10.xlsx"

# Read the excell file and sheet for input
wb = openpyxl.load_workbook('All in one Sup Conf 10%.xlsx')
sheet_input = wb['Sheet1']

# Read the excell file and sheet for output
wb2 = openpyxl.Workbook()
sheet_output = wb2.active

for row in sheet_input['f2':'f19769']:  # read perticular cell not all cell in the sheet
    try:
        for cell in row:
            x[i] = cell.value  # put the cell value into the list x[i]
            word = x[i]
            print(word)

            word2 = stemSentence(word)
            print(word2)

            word3 = word2.split()
            word3 = sorted(set(word3))
            word3 = str(word3)
            word3 = word3.replace('[', '')
            word3 = word3.replace(']', '')
            word3 = word3.replace('\'', '')
            word3 = word3.replace(',', '')
            word3 = word3.replace('.', ' ')
            word3 = word3.replace(';', ' ')
            word3 = word3.replace('-PRON- ', '')
            word3 = word3.replace('  ', ' ')

            print(word3)

            a2 = sheet_output.cell(row=row_no, column=3)
            a2.value = word3
            row_no = row_no + 1

            wb2.save(z)
        wb2.save(z)
    except:
         row_no = row_no + 1
         pass
         wb2.save(z)
