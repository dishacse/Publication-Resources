///
Author: Md. Shamsujjoha, Prof. John Grundy, Dr. Li Li, Dr. Qinghua Lu and Dr. Hourieh Khalajzade
Python Version: 3.7 or later
///

import re
import os
import fnmatch
import glob
import openpyxl
import Remove Java Keywords

row_no = 1
wb = openpyxl.Workbook()
sheet = wb.active

# z_name = 1
myFiles = ['001997E04F504112B5E018DD86261CFE47884E3EE1CEABB14AFCB4B8C122C6AB'] // this is an example app, during the computation we have use 5000 apps at a time where SHA256 are seperated by commas

for filename in myFiles:
    name = filename
    name2='=IF(ISBLANK(C',row_no,'),0,LEN(TRIM(C',row_no,'))-LEN(SUBSTITUTE(C',row_no,',\" \",\"\"))+1)'
    name2 = str(name2)

    folder_path = os.path.join('D1', filename) // we assume our example is inside D1 directory
    f2 = open('output1.txt', 'w+', encoding="utf8")

    for dirpath, dirs, files in os.walk(folder_path):
        for filename in fnmatch.filter(files, '*.smali'):
            f = open(os.path.join(dirpath, filename), encoding="utf8")
            for word in f:
                i = 8
                if re.match("(.*)\.method(.*)", word):
                    try:
                        while (i < len(word)):
                            if (word[i] is not '('):
                                print(word[i], end="", file = f2)
                                i = i + 1
                            else:
                                print("", file = f2),
                                break
                    except IndexError:
                        i = 8

    f2.close()

    # This user define function removes all Java keyword like public, static
    # from output1.txt file and store it in output2.txt file
    Java11.removeJavaKeyword()
    ###################################################################################################

    f5 = open('output2.txt', 'r+', encoding="utf8")
    f6 = open('output3.txt', 'w+', encoding="utf8")

    for word3 in f5:
        i = 0
        try:
            while (i < len(word3)):
                if (word3[i] == '\n'):
                    i = i + 1
                if (word3[i].isupper()):
                    print(" ", end="", file=f6)
                print(word3[i], end="", file=f6)
                i = i + 1
        except IndexError:
            i = 0
            print(" ", end="", file=f6)
            
    f5.close()
    f6.close()
    print("\n")

    f7 = open('output3.txt', 'r+', encoding="utf8")
    f8 = open('output4.txt', 'w+', encoding="utf8")

    for word4 in f7:
        i = 0
        try:
            while (i < len(word4)):
                if ((word4[i] == ' ' and word4[i + 1] <= ' ')):
                    i = i + 1
                elif (word4[i] >= '0' and word4[i] <= '9'):
                    i = i + 1
                elif ((word4[i] == ' ' and ((word4[i + 1] >= 'a' and word4[i + 1] <= 'z') or (
                        word4[i + 1] >= 'A' and word4[i + 1] <= 'Z')) and word4[i + 2] <= ' ')):
                    i = i + 2
                else:
                    print(word4[i], end="", file=f8)
                    i = i + 1
        except IndexError:
            i = 0

    f7.close()
    f8.close()

    f9 = open('output4.txt', 'r+', encoding="utf8")

    read_final_output_to_write_it_in_excell = f9.read()

    c2 = sheet.cell(row=row_no, column=1)
    c2.value = name

    c3 = sheet.cell(row=row_no, column=2)
    c3.value = name2

    c4 = sheet.cell(row=row_no, column=3)
    c4.value = read_final_output_to_write_it_in_excell

    row_no = row_no + 1
    wb.save("D1.xlsx")

wb.save("D1.xlsx")